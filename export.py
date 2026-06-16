from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "outputs"


def add_chart_sheets(year_start, year_end):
    wb = load_workbook(BASE_DIR / "energy_data.xlsx")

    for sheet_name in ["Electricity Mix", "Total Energy Mix", "Solar vs Wind", "GDP vs Renewables"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    def add_chart_sheet(sheet_name, title, image_filename):
        ws = wb.create_sheet(sheet_name)

        ws["A1"] = title
        ws.merge_cells("A1:F1")
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].fill = PatternFill(fill_type="solid", start_color="FFD700")
        ws["A1"].alignment = Alignment(horizontal="center")

        img = Image(OUTPUT_DIR / image_filename)

        ws.add_image(img, "A4")

    add_chart_sheet("Electricity Mix",
                    f"Renewables Share of Electricity {year_start}-{year_end}",
                    "renewables_share_elec.png")

    add_chart_sheet("Total Energy Mix",
                    f"Renewable Share of Total Energy ({year_start}-{year_end})",
                    "renewables_share_energy.png")

    add_chart_sheet("Solar vs Wind",
                    f"Solar and Wind Electricity Share ({year_start}-{year_end})",
                    "solar_vs_wind.png")

    add_chart_sheet("GDP vs Renewables",
                    f"Renewables vs GDP per Capita ({year_start}-{year_end})",
                    "gdp_vs_renewables.png")

    wb.save(BASE_DIR / "energy_data.xlsx")
